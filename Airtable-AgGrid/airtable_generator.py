import calendar
import time
from datetime import timedelta

from aitable_helper import *
import os
from openpyxl import Workbook, load_workbook

s_requests_get = "get"
endpoint_url, endpoint_token = "https://productionv2.qikpod.com:8989/", "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJleHAiOjE3NTM5NjE5MzV9.VLBNCnYaTSNpT6D7-uw9PABnHL-2MZwNkld5a-DdNPc"
month = get_now().strftime("%b")
file_path = f"excel_data/{month}_Excel_Data.xlsx"

bRunNow_A = bRunNow_B = bRunNow_C = None
last_run_date_A = last_run_date_B = last_run_date_C = None


def create_workbook(excel_path: str, sheet_name: str, column_fields: list, delete_flag: bool):
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_index = wb.sheetnames.index(sheet_name) if sheet_name in wb.sheetnames else None
    if delete_flag and sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"sheet_name: {sheet_name} deleted...!")

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name, index=sheet_index) if sheet_index is not None else wb.create_sheet(title=sheet_name)
        for col_idx, header in enumerate(column_fields, start=1):
            ws.cell(row=1, column=col_idx, value=header)
    else:
        ws = wb[sheet_name]
    print({"sheet_name": sheet_name, "delete_flag": delete_flag})
    wb.save(excel_path)
    return wb, ws



def insert_records_in_workbook(ws, column_fields: list, record_generator):
    header_row = [cell.value for cell in ws[1]]
    if header_row != column_fields:
        return {"status": "failure", "message": "Column fields do not match with existing headers"}
    for record in record_generator:
        ws.append([record.get(field, "") for field in column_fields])
    print({"status": "success", "file_path": f"{file_path}", "message": "Records inserted"}, "\n" )
    return {"status": "success", "file_path": f"{file_path}", "message": "Records inserted"}


def upsert_record_in_workbook(ws, column_fields: list, id_field: str, record_generator):
    header_row = [cell.value for cell in ws[1]]
    if header_row != column_fields:
        print({"status": "failure", "message": "Column fields do not match with existing headers"})
        return {"status": "failure", "message": "Column fields do not match with existing headers"}
    try:
        id_col_idx = column_fields.index(id_field) + 1  # 1-based index for Excel columns
    except ValueError:
        print({"status": "failure", "message": f"'{id_field}' not found in column headers"})
        return {"status": "failure", "message": f"'{id_field}' not found in column headers"}
    for record in record_generator:

        record_id = str(record.get(id_field))
        updated = False  # Flag to check if record was updated
        for row in ws.iter_rows(min_row=2, max_col=len(column_fields)):
            if str(row[id_col_idx - 1].value) == record_id:
                for field, value in record.items():
                    if field in column_fields:
                        row[column_fields.index(field)].value = value
                updated = True
                break  # Stop after updating the first match

        if not updated:
            ws.append([record.get(field, "") for field in column_fields])

    print( {"status": "success", "file_path": f"{file_path}", "message": "Records processed (updated/inserted)"}, "\n" )
    return {"status": "success", "file_path": f"{file_path}", "message": "Records processed (updated/inserted)"}


def get_pinged_at_in_minutes(pinged_at_raw):
    new_flag = {"alert": "🚨 ALERT", "ok": "🟢 OK"}
    if not pinged_at_raw:
        return f"{new_flag['alert']}", ""

    ist = timezone("Asia/Kolkata")
    now = get_now()
    try:
        pinged_at_dt = datetime.datetime.strptime(pinged_at_raw, "%Y-%m-%d %H:%M:%S.%f")
        pinged_at_dt = ist.localize(pinged_at_dt)
        diff_minutes = int((now - pinged_at_dt).total_seconds() / 60)

        status = new_flag["alert"] if diff_minutes > 15 else new_flag["ok"]
        return status, diff_minutes

    except Exception:
        print("Failed to parse:", pinged_at_raw)
        return new_flag['alert'], pinged_at_raw

def get_fe_monitor_data(column_fields: list):
    pod_response = pytest_call_rest_api(request_type=s_requests_get, endpoint=endpoint_url + 'pods/', token=endpoint_token)
    if pod_response.get("status") == "success":
        for pod in pod_response["records"]:
            if pod.get("status") != "active":
                continue
            loc_id = pod.get("location_id")
            loc_response = pytest_call_rest_api(request_type=s_requests_get, endpoint=endpoint_url + f'locations/{loc_id}', token=endpoint_token)
            for loc in loc_response["records"]:
                new_flag, pinged_at = get_pinged_at_in_minutes(pod.get("pinged_at"))
                data = {
                    "pod_id": pod.get("id"), "location_name": loc.get("location_name"), "location_id": loc_id, "new_flag": new_flag,
                    "primary_fe": loc.get("primary_fe"), "pod_name": pod.get("pod_name"), "pinged_at (Min)": pinged_at,
                    "pod_power_status": pod.get("pod_power_status"), "updated_at": pod.get("updated_at"),
                    "fe_tag": pod.get("fe_tag") or "None", "pod_mode": pod.get("pod_mode"), "pod_connection_method": pod.get("pod_connection_method"),
                    "last modified": str(get_now())
                }
                yield {field: data.get(field, "") for field in column_fields}


def get_pod_frequency_data():
    now = get_now()
    current_day = now.day
    fifteen_minutes_ago = now - timedelta(minutes=15)
    locations = pytest_call_rest_api(request_type=s_requests_get, endpoint=endpoint_url + 'locations/', token=endpoint_token)
    for location in locations["records"]:
        pods_response = pytest_call_rest_api(request_type=s_requests_get, endpoint=endpoint_url + 'pods/', params={"location_id": location.get('id')}, token=endpoint_token)
        if pods_response.get("status") != "success":
            continue
        for pod in pods_response["records"]:
            if pod.get("status") != "active":
                continue
            ping_at = pod.get("pinged_at")
            power_status = pod.get("pod_power_status")
            if str(fifteen_minutes_ago) > str(ping_at) or power_status == 'UPS':
                yield { "location_name": location.get("location_name"), "location_id": str(location.get('id')), "primary_fe": location.get('primary_fe'), f'day{current_day}': power_status ,"last modified": str(now)}
                break  # only one pod per location to be considered
    return


def get_fe_evaluation_data(location_id, server_endpoint):
    ist = timezone("Asia/Kolkata")
    date = get_now()  # already IST-aware
    year = date.year
    month = date.month # current Month May It will return 5
    current_day = date.day #current Day It will return day7
    first_of_month = ist.localize(datetime.datetime(year, month, 1, 0, 0, 0))
    end_of_month = ist.localize(datetime.datetime(year, month, current_day, 23, 59, 59, 999999))

    start_of_day = ist.localize(datetime.datetime(year, month, current_day, 0, 0, 0))
    end_of_day = ist.localize(datetime.datetime(year, month, current_day, 23, 59, 59, 999999))
    # print(f"first_of_month: {first_of_month}, end_of_month: {end_of_month}, current_day: {current_day}, start_of_day: {start_of_day}, end_of_day: {end_of_day}")

    today_count, month_total = 0, 0
    params = {"location_id": location_id}
    response = pytest_call_rest_api(request_type=s_requests_get, endpoint=endpoint_url + server_endpoint,params=params, token=endpoint_token)
    if response["status"] == "success":
        for item in response["records"]:
            created_at = datetime.datetime.strptime(item["created_at"], "%Y-%m-%dT%H:%M:%S.%f")
            created_at = ist.localize(created_at)
            if item["reservation_status"] in ["PickupPending", "PickupCompleted"]:
                if start_of_day <= created_at <= end_of_day:
                    today_count += 1
                if first_of_month <= created_at <= end_of_month:
                    month_total += 1

    return today_count, month_total



def run_fe_report_data():
    now = get_now()
    current_day = now.day
    response = pytest_call_rest_api(request_type=s_requests_get, endpoint=endpoint_url + 'locations/', token=endpoint_token)
    if response["status"] == "success":
        for item in response["records"]:
            locationID = item["id"]
            adhoc_today, adhoc_total = get_fe_evaluation_data(location_id=locationID, server_endpoint="adhoc/reservations/")
            standard_today, standard_total = get_fe_evaluation_data(location_id=locationID, server_endpoint="reservations/")
            average_per_day = round((adhoc_total + standard_total) / current_day)
            data = { "location_name": item.get("location_name"), "location_id": locationID, "primary_fe": item.get("primary_fe"), "primary_bd": item.get('primary_bd'), "secondary_fe": item.get("secondary_fe"),
                    "today_count": adhoc_today+standard_today, "total_count": adhoc_total+standard_total, "Average Per day": average_per_day, "last modified": str(now)}

            yield data

def get_reservation_data():
    now = get_now()
    response = pytest_call_rest_api( request_type=s_requests_get, endpoint=endpoint_url + 'locations/', token=endpoint_token )
    if response["status"] == "success":
        for record in response["records"]:
            location_id = record["id"]
            location_name = record["location_name"]
            primary_fe = record.get('primary_fe')
            reservation_counts = {"location_name": location_name, "location_id": location_id,"primary_fe": primary_fe,"24Hrs": 0,"48Hrs": 0,"7Days": 0,"30Days": 0,"Pickup_Pending": 0,"RTOCompleted": 0,"Pickup_Completed": 0,"Total_Count": 0,"last modified": str(now) }
            for server_endpoint in ["adhoc/reservations/", "reservations/"]:
                record_dict = get_reservation_count(location_id, server_endpoint)
                for key in record_dict:
                    if key != "location_id":
                        reservation_counts[key] += record_dict[key]
            yield reservation_counts


def get_reservation_count(location_id, server_endpoint):
    ist = timezone("Asia/Kolkata")
    now = get_now()
    reservation_counts = { "location_id": location_id, "24Hrs": 0, "48Hrs": 0, "7Days": 0, "30Days": 0, "Pickup_Pending": 0, "RTOCompleted": 0, "Pickup_Completed": 0, "Total_Count": 0 }
    params = {"location_id": location_id}
    response = pytest_call_rest_api( request_type=s_requests_get, endpoint=endpoint_url + server_endpoint, params=params, token=endpoint_token )
    if response["status"] == "success":
        reservation_counts["Total_Count"] = response.get("count", 0)
        for record in response["records"]:
            created_at = datetime.datetime.strptime(record["created_at"], "%Y-%m-%dT%H:%M:%S.%f")
            created_time = ist.localize(created_at)

            days_difference = (now - created_time).days

            if days_difference == 0:
                reservation_counts["24Hrs"] += 1
            elif days_difference == 1:
                reservation_counts["48Hrs"] += 1
            elif 1 < days_difference <= 7:
                reservation_counts["7Days"] += 1
            elif 7 < days_difference <= 30:
                reservation_counts["30Days"] += 1

            status = record.get("reservation_status")
            if status == "PickupPending":
                reservation_counts["Pickup_Pending"] += 1
            elif status == "RTOCompleted":
                reservation_counts["RTOCompleted"] += 1
            elif status == "PickupCompleted":
                reservation_counts["Pickup_Completed"] += 1

    return reservation_counts



def run_fe_monitor_report(sheet_name: str, delete_sheet: bool = False):
    custom_fields = ["pod_id", "location_name", "location_id", "new_flag", "updated_at", "primary_fe", "pod_name", "pinged_at (Min)", "pod_power_status", "fe_tag", "pod_mode", "pod_connection_method", "last modified"]
    wb, ws = create_workbook(excel_path=file_path, sheet_name=sheet_name, column_fields=custom_fields, delete_flag=delete_sheet)
    result = insert_records_in_workbook(ws=ws, column_fields=custom_fields, record_generator=get_fe_monitor_data(column_fields=custom_fields))
    wb.save(file_path)
    # return result


def run_pod_frequency_report(sheet_name: str, delete_sheet: bool = False):
    custom_fields = ['location_name', 'location_id', 'primary_fe'] + [f'day{i}' for i in range(1, 32)] + ["last modified"]
    wb, ws = create_workbook(excel_path=file_path, sheet_name=sheet_name, column_fields=custom_fields, delete_flag=delete_sheet)
    result = upsert_record_in_workbook(ws, custom_fields, id_field="location_id", record_generator=get_pod_frequency_data())
    wb.save(file_path)
    return result

def run_fe_evaluation_report(sheet_name: str, delete_sheet: bool = False):
    custom_fields = ['location_name', 'location_id', "primary_fe", "primary_bd", "today_count", "total_count", "Average Per day", "last modified"]
    wb, ws = create_workbook(excel_path=file_path, sheet_name=sheet_name, column_fields=custom_fields, delete_flag=delete_sheet)
    result = insert_records_in_workbook(ws=ws, column_fields=custom_fields, record_generator=run_fe_report_data())
    wb.save(file_path)
    return result


def run_reservation_report(sheet_name: str, delete_sheet: bool = False):
    custom_fields = ['location_name', 'location_id', "primary_fe", "24Hrs", "48Hrs", "7Days", "30Days", "Pickup_Pending", "RTOCompleted", "Pickup_Completed", "Total_Count", "last modified"]
    wb, ws = create_workbook(excel_path=file_path, sheet_name=sheet_name, column_fields=custom_fields, delete_flag=delete_sheet)
    result = insert_records_in_workbook(ws=ws, column_fields=custom_fields, record_generator=get_reservation_data())
    wb.save(file_path)
    # return result

def airtabletasks_A() -> int:
    global bRunNow_A, last_run_date_A
    current_time = get_now()
    try:
        if last_run_date_A is None or (current_time - last_run_date_A) >= timedelta(minutes=10):
            bRunNow_A = True

        if bRunNow_A:
            print(f"airtabletasks_A is started: [ {current_time} ] ")
            run_fe_monitor_report(sheet_name="Fe Monitor Report", delete_sheet=True)
            bRunNow_A = False
            last_run_date_A = current_time
        return 1
    except Exception as e:
        print(f"An exception occurred: {str(e)}")
    return 0


def airtabletasks_B() -> int:
    global bRunNow_B, last_run_date_B
    try:
        current_time = get_now()
        if ((last_run_date_B is None) or ((current_time.hour != last_run_date_B.hour)) and (current_time.minute >= 50)):
            bRunNow_B = True

        if bRunNow_B:
            print(f"airtabletasks_B is started: [ {current_time} ] ")
            run_fe_evaluation_report(sheet_name="Evaluation Fe Report", delete_sheet=True)
            run_reservation_report(sheet_name="Reservation Report", delete_sheet=True)
            bRunNow_B = False
            last_run_date_B = current_time
        return 1
    except Exception as e:
        print(f"An exception occurred: {str(e)}")
    return 0




def airtabletasks_C() -> int:
    global bRunNow_C, last_run_date_C
    try:
        current_time = get_now()
        if ((last_run_date_C is None) or (current_time.day != last_run_date_C.day)) and (current_time.hour >= 16):
            bRunNow_C = True

        if bRunNow_C:
            print(f"airtabletasks_C is started: [ {current_time} ] ")
            run_pod_frequency_report(sheet_name="Pod Frequency Report", delete_sheet=False)
            last_run_date_C = current_time
            bRunNow_C = False
        return 1
    except Exception as e:
        print(f"An exception occurred: {str(e)}")
    return 0



if __name__ == "__main__":
    while True:
        print(f"********************* AirtableTask Cron Started At: {get_now()} ************************\n")
        airtabletasks_A()
        # airtabletasks_B()
        # airtabletasks_C()
        print(f"********************* AirtableTask Cron Ended At: {get_now()} ************************\n")
        time.sleep(45)
